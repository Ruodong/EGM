"""Ask EGM — AI-assisted review analysis (OpenAI-compatible LLM + RAG).

Enhanced with:
  - Image & file attachment support (multimodal vision)
  - Structured markdown responses (tables, lists, code blocks)
  - Follow-up question recommendations after each response
"""

from __future__ import annotations

import asyncio
import base64
import io
import json
import logging
import re
from typing import AsyncGenerator

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import require_permission, get_current_user, AuthUser, Role
from app.config import settings
from app.database import get_db
from app.services.tavily_search import (
    is_tavily_configured, tavily_search, format_search_results_for_llm, WEB_SEARCH_TOOL,
)
from app.utils.embeddings import find_similar_reviews, upsert_review_embedding

logger = logging.getLogger(__name__)


# ── Object-level authorization ────────────────────────────────────────────────

async def _check_review_access(
    db: AsyncSession, user: AuthUser, domain_review_id: str
) -> dict:
    """Verify user has access to this domain review's AI chat features.

    - Admin / Governance Lead: always allowed
    - Domain Reviewer: only if the review's domain_code is in user.domain_codes
    - Requestor: only if they own the governance request

    Returns the review lookup dict.  Raises 404 / 403 on failure.
    """
    row = (await db.execute(text("""
        SELECT dr.id, dr.domain_code, dr.request_id, gr.requestor
        FROM domain_review dr
        JOIN governance_request gr ON dr.request_id = gr.id
        WHERE dr.id = :rid
    """), {"rid": domain_review_id})).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Domain review not found")

    r = dict(row)

    if Role.ADMIN in user.roles:
        return r
    if Role.GOVERNANCE_LEAD in user.roles:
        return r
    if Role.DOMAIN_REVIEWER in user.roles and r["domain_code"] in (user.domain_codes or []):
        return r
    if Role.REQUESTOR in user.roles and r["requestor"] == user.id:
        return r

    raise HTTPException(
        status_code=403,
        detail="Access denied: you don't have permission for this domain review",
    )

router = APIRouter()

# Max file size: 10 MB
MAX_FILE_SIZE = 10 * 1024 * 1024
ALLOWED_IMAGE_TYPES = {"image/png", "image/jpeg", "image/gif", "image/webp"}
ALLOWED_FILE_TYPES = ALLOWED_IMAGE_TYPES | {
    "application/pdf",
    "text/plain",
    "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}

# Max characters of extracted text to include in LLM context
_MAX_EXTRACTED_TEXT = 8000


def _extract_text_from_file(data: bytes, content_type: str) -> str | None:
    """Extract readable text from PDF, DOCX, or XLSX binary data.

    Returns extracted text (truncated to _MAX_EXTRACTED_TEXT) or None on failure.
    """
    try:
        if content_type == "application/pdf":
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(data))
            pages = []
            for i, page in enumerate(reader.pages):
                t = page.extract_text()
                if t:
                    pages.append(f"--- Page {i + 1} ---\n{t}")
                if sum(len(p) for p in pages) > _MAX_EXTRACTED_TEXT:
                    break
            return "\n".join(pages)[:_MAX_EXTRACTED_TEXT] if pages else None

        if content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            from docx import Document

            doc = Document(io.BytesIO(data))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            text = "\n".join(paragraphs)
            return text[:_MAX_EXTRACTED_TEXT] if text else None

        if content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets:
                parts.append(f"--- Sheet: {ws.title} ---")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    parts.append("\t".join(cells))
                if sum(len(p) for p in parts) > _MAX_EXTRACTED_TEXT:
                    break
            wb.close()
            text = "\n".join(parts)
            return text[:_MAX_EXTRACTED_TEXT] if text else None

    except Exception as e:
        logger.warning("Text extraction failed for %s: %s", content_type, e)
    return None


# ── LLM configuration ───────────────────────────────────────────────────────


def _get_openai_client():
    """Lazy-init async OpenAI-compatible client."""
    from openai import AsyncOpenAI
    if not settings.LLM_BASE_URL or not settings.LLM_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="LLM is not configured. Set LLM_BASE_URL and LLM_API_KEY environment variables.",
        )
    return AsyncOpenAI(
        base_url=settings.LLM_BASE_URL,
        api_key=settings.LLM_API_KEY,
    )


# ── Follow-up question extraction ───────────────────────────────────────────

FOLLOW_UP_MARKER = "[FOLLOW_UP]"

def _extract_follow_ups(text_content: str) -> tuple[str, list[str]]:
    """Extract follow-up questions from the end of an LLM response.

    Returns (cleaned_content, follow_up_questions).
    """
    if FOLLOW_UP_MARKER not in text_content:
        return text_content, []

    idx = text_content.rfind(FOLLOW_UP_MARKER)
    main_content = text_content[:idx].rstrip()
    follow_up_block = text_content[idx + len(FOLLOW_UP_MARKER):].strip()

    questions: list[str] = []
    for line in follow_up_block.split("\n"):
        line = line.strip()
        # Match numbered or bulleted lines: "1. ...", "- ...", "• ..."
        m = re.match(r'^(?:\d+[\.\)]\s*|[-•]\s*)(.*)', line)
        if m and m.group(1).strip():
            questions.append(m.group(1).strip())
        elif line and not line.startswith("["):
            questions.append(line)

    return main_content, questions[:5]  # max 5


# ── Helpers ──────────────────────────────────────────────────────────────────

async def _build_system_prompt(db: AsyncSession, domain_review_id: str, *, web_search: bool = False) -> str:
    """Collect all context for the domain review and build a system prompt."""

    # 1. Get domain review + request info
    review_row = (await db.execute(text("""
        SELECT dr.id, dr.domain_code, dr.status, dr.outcome, dr.reviewer, dr.reviewer_name,
               gr.id AS request_uuid, gr.request_id, gr.title AS gov_title,
               gr.project_name, gr.project_code, gr.project_proj_type,
               gr.project_description, gr.project_pm, gr.project_pm_itcode,
               gr.project_dt_lead, gr.project_it_lead,
               gr.project_start_date, gr.project_go_live_date,
               gr.product_software_type, gr.product_end_user, gr.user_region,
               gr.third_party_vendor,
               gr.requestor, gr.requestor_name, gr.business_unit,
               gr.gov_project_type, gr.description AS gov_description,
               dn.domain_name AS domain_name
        FROM domain_review dr
        JOIN governance_request gr ON dr.request_id = gr.id
        LEFT JOIN domain_registry dn ON dr.domain_code = dn.domain_code
        WHERE dr.id = :rid
    """), {"rid": domain_review_id})).mappings().first()

    if not review_row:
        raise HTTPException(status_code=404, detail="Domain review not found")

    r = dict(review_row)
    request_uuid = str(r["request_uuid"])
    domain_code = r["domain_code"]

    # 2. Questionnaire responses (requestor answers for this domain)
    q_rows = (await db.execute(text("""
        SELECT t.section, t.question_no, t.question_text, t.answer_type,
               rqr.answer
        FROM request_questionnaire_response rqr
        JOIN domain_questionnaire_template t ON rqr.template_id = t.id
        WHERE rqr.request_id = :req_id
          AND rqr.domain_code = :dc
          AND t.is_active = TRUE
        ORDER BY t.sort_order, t.question_no
    """), {"req_id": request_uuid, "dc": domain_code})).mappings().all()

    # 3. Action items + feedback
    action_rows = (await db.execute(text("""
        SELECT a.action_no, a.title, a.description, a.priority, a.action_type,
               a.status, a.assignee_name, a.id AS action_id
        FROM review_action a
        WHERE a.domain_review_id = :rid
        ORDER BY a.action_no
    """), {"rid": domain_review_id})).mappings().all()

    feedback_by_action: dict[str, list[dict]] = {}
    if action_rows:
        action_ids = [str(a["action_id"]) for a in action_rows]
        fb_rows = (await db.execute(text("""
            SELECT action_id, round_no, feedback_type, content, created_by_name
            FROM review_action_feedback
            WHERE action_id = ANY(:ids)
            ORDER BY create_at
        """), {"ids": action_ids})).mappings().all()
        for fb in fb_rows:
            aid = str(fb["action_id"])
            feedback_by_action.setdefault(aid, []).append(dict(fb))

    # ── Build prompt ─────────────────────────────────────────────────────

    parts: list[str] = []
    parts.append(
        'You are "Ask EGM", an AI assistant helping governance reviewers analyze compliance review requests.\n'
        "You MUST ONLY answer questions directly related to the current domain review context provided below.\n"
        "If the user asks about topics unrelated to this governance review, politely decline and suggest they ask about this review instead.\n"
        "Provide thorough, professional analysis. Answer in the same language as the user's question.\n"
        "\n"
        "## Response Format Guidelines\n"
        "- Use **Markdown** formatting in your responses for clarity.\n"
        "- When comparing data or listing multiple items, use **tables** (Markdown table syntax).\n"
        "- Use **bold** for key terms, **bullet lists** for enumerations, and **headings** (##, ###) for sections.\n"
        "- When the user uploads an image or file, analyze its content in relation to this review.\n"
        "\n"
        "## Follow-up Questions\n"
        "At the END of every response, suggest 2-3 relevant follow-up questions the reviewer might want to ask.\n"
        "Format them exactly like this (the marker must be on its own line):\n"
        "\n"
        "[FOLLOW_UP]\n"
        "1. First suggested question?\n"
        "2. Second suggested question?\n"
        "3. Third suggested question?\n"
    )

    if web_search:
        parts.append(
            "## Web Search\n"
            "You have access to a `web_search` tool for looking up external information.\n"
            "Use it when the user asks about regulations, standards, vendors, technologies, "
            "or anything requiring up-to-date knowledge beyond this review's context.\n"
            "When citing web search results, use inline markdown links: [Source Title](url).\n"
        )

    # Project / Request info
    parts.append("## Current Review Context")
    parts.append(f"- **Domain**: {r.get('domain_name') or domain_code}")
    parts.append(f"- **Review Status**: {r.get('status')}")
    if r.get("outcome"):
        parts.append(f"- **Outcome**: {r['outcome']}")
    parts.append(f"- **Reviewer**: {r.get('reviewer_name') or r.get('reviewer') or 'N/A'}")
    parts.append("")

    parts.append("## Project Information")
    parts.append(f"- **Request ID**: {r.get('request_id')}")
    parts.append(f"- **Title**: {r.get('gov_title') or 'N/A'}")
    parts.append(f"- **Description**: {r.get('gov_description') or 'N/A'}")
    parts.append(f"- **Project Name**: {r.get('project_name') or 'N/A'}")
    parts.append(f"- **Project Type**: {r.get('project_proj_type') or r.get('gov_project_type') or 'N/A'}")
    parts.append(f"- **Business Unit**: {r.get('business_unit') or 'N/A'}")
    parts.append(f"- **Requestor**: {r.get('requestor_name') or r.get('requestor')}")
    parts.append(f"- **Project Manager**: {r.get('project_pm') or 'N/A'}")
    if r.get("product_software_type"):
        parts.append(f"- **Software Type**: {r['product_software_type']}")
    if r.get("third_party_vendor"):
        parts.append(f"- **Third Party Vendor**: {r['third_party_vendor']}")
    if r.get("product_end_user"):
        parts.append(f"- **End Users**: {', '.join(r['product_end_user']) if isinstance(r['product_end_user'], list) else r['product_end_user']}")
    if r.get("user_region"):
        parts.append(f"- **Regions**: {', '.join(r['user_region']) if isinstance(r['user_region'], list) else r['user_region']}")
    parts.append("")

    # Questionnaire
    if q_rows:
        parts.append("## Questionnaire Responses")
        current_section = None
        for q in q_rows:
            section = q.get("section") or "General"
            if section != current_section:
                parts.append(f"\n### {section}")
                current_section = section
            answer_data = q.get("answer")
            if isinstance(answer_data, str):
                try:
                    answer_data = json.loads(answer_data)
                except (json.JSONDecodeError, TypeError):
                    pass
            if isinstance(answer_data, dict):
                val = answer_data.get("value", "")
                if isinstance(val, list):
                    val = ", ".join(val)
                other = answer_data.get("otherText", "")
                desc = answer_data.get("descriptionText", "")
                answer_str = str(val or "(no answer)")
                if other:
                    answer_str += f" (Other: {other})"
                if desc:
                    answer_str += f" — {desc}"
            else:
                answer_str = str(answer_data) if answer_data else "(no answer)"
            parts.append(f"**Q{q['question_no']}. {q['question_text']}**")
            parts.append(f"A: {answer_str}")
        parts.append("")

    # Action items
    if action_rows:
        parts.append("## Action Items")
        for a in action_rows:
            aid = str(a["action_id"])
            parts.append(f"\n### Action #{a.get('action_no', '?')}: {a['title']} [{a['status']}]")
            parts.append(f"- Priority: {a['priority']}, Type: {a['action_type']}")
            if a.get("description"):
                parts.append(f"- Description: {a['description']}")
            if a.get("assignee_name"):
                parts.append(f"- Assignee: {a['assignee_name']}")
            fbs = feedback_by_action.get(aid, [])
            if fbs:
                parts.append("- Feedback:")
                for fb in fbs:
                    role_label = "Reviewer" if fb["feedback_type"] == "follow_up" else "Assignee"
                    parts.append(f"  - [{role_label}] {fb['content']}")
        parts.append("")

    # 4. Similar historical reviews (RAG)
    try:
        current_summary_parts = [
            r.get("domain_name") or domain_code,
            r.get("gov_title") or "",
            r.get("gov_description") or "",
            r.get("project_proj_type") or "",
            r.get("product_software_type") or "",
        ]
        query_text = " ".join(p for p in current_summary_parts if p)
        similar = await find_similar_reviews(
            db, domain_review_id, domain_code, query_text, top_k=5
        )
        if similar:
            parts.append("## Similar Historical Reviews in This Domain")
            parts.append("The following are past reviews in the same domain that may be relevant for reference:\n")
            for i, s in enumerate(similar, 1):
                parts.append(f"### Similar Case #{i} (similarity: {s['similarity']})")
                summary_preview = s["content_summary"][:500]
                if len(s["content_summary"]) > 500:
                    summary_preview += "..."
                parts.append(summary_preview)
                parts.append("")
    except Exception as e:
        logger.warning("RAG similar review lookup failed (non-fatal): %s", e)

    return "\n".join(parts)


def _build_multimodal_content(
    text_content: str, attachments: list[dict]
) -> list[dict]:
    """Build OpenAI multimodal content array with text + images.

    For images: include as base64 data URL for vision.
    For non-image files: include text description of the file.
    """
    content_parts: list[dict] = []

    # Text part first
    if text_content:
        content_parts.append({"type": "text", "text": text_content})

    for att in attachments:
        ct = att.get("contentType", "")
        if ct in ALLOWED_IMAGE_TYPES:
            # Image → base64 vision
            b64 = base64.b64encode(att["data"]).decode("utf-8")
            content_parts.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:{ct};base64,{b64}",
                    "detail": "auto",
                },
            })
        else:
            # Non-image file → extract text and include as context
            fname = att.get("fileName", "unknown")
            fsize = att.get("fileSize", 0)
            desc = f"[Attached file: {fname} ({fsize} bytes, type: {ct})]"
            # Try to extract text content
            if ct in ("text/plain", "text/csv"):
                try:
                    file_text = att["data"].decode("utf-8", errors="replace")[:_MAX_EXTRACTED_TEXT]
                    desc += f"\n\nFile contents:\n```\n{file_text}\n```"
                except Exception:
                    pass
            elif ct in (
                "application/pdf",
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ):
                extracted = _extract_text_from_file(att["data"], ct)
                if extracted:
                    desc += f"\n\nExtracted contents:\n```\n{extracted}\n```"
                else:
                    desc += "\n\n(Could not extract text content from this file.)"
            content_parts.append({"type": "text", "text": desc})

    return content_parts


# ── Request / response models ────────────────────────────────────────────────

class ChatRequest(BaseModel):
    message: str
    attachmentIds: list[str] | None = None
    webSearch: bool = False


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.post(
    "/{domain_review_id}/upload",
    dependencies=[Depends(require_permission("review_action", "read"))],
)
async def upload_attachment(
    domain_review_id: str,
    file: UploadFile,
    user: AuthUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Upload a file attachment for chat (images, documents)."""
    await _check_review_access(db, user, domain_review_id)
    content_type = file.content_type or "application/octet-stream"
    if content_type not in ALLOWED_FILE_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"File type '{content_type}' not supported. Allowed: images, PDF, TXT, CSV, XLSX, DOCX.",
        )

    data = await file.read()
    if len(data) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 10 MB.")

    row = (await db.execute(text("""
        INSERT INTO ask_egm_attachment
            (domain_review_id, file_name, file_size, content_type, file_data, create_by)
        VALUES (:rid, :fname, :fsize, :ctype, :fdata, :uid)
        RETURNING id, file_name, file_size, content_type, create_at
    """), {
        "rid": domain_review_id,
        "fname": file.filename or "untitled",
        "fsize": len(data),
        "ctype": content_type,
        "fdata": data,
        "uid": user.id,
    })).mappings().first()
    await db.commit()

    return {
        "id": str(row["id"]),
        "fileName": row["file_name"],
        "fileSize": row["file_size"],
        "contentType": row["content_type"],
        "isImage": content_type in ALLOWED_IMAGE_TYPES,
        "createAt": row["create_at"].isoformat() if row["create_at"] else None,
    }


@router.get(
    "/attachments/{att_id}",
    dependencies=[Depends(require_permission("review_action", "read"))],
)
async def download_attachment(
    att_id: str,
    user: AuthUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download a chat attachment (returns binary)."""
    row = (await db.execute(text("""
        SELECT a.file_name, a.content_type, a.file_data, a.domain_review_id
        FROM ask_egm_attachment a WHERE a.id = :aid
    """), {"aid": att_id})).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found")
    # Object-level access check via the attachment's parent domain review
    await _check_review_access(db, user, str(row["domain_review_id"]))
    disposition = "inline" if row["content_type"].startswith("image/") else "attachment"
    return Response(
        content=bytes(row["file_data"]),
        media_type=row["content_type"],
        headers={"Content-Disposition": f'{disposition}; filename="{row["file_name"]}"'},
    )


@router.delete(
    "/attachments/{att_id}",
    dependencies=[Depends(require_permission("review_action", "read"))],
)
async def delete_attachment(
    att_id: str,
    user: AuthUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete a single attachment (e.g. user removed it before sending)."""
    row = (await db.execute(text(
        "SELECT domain_review_id, create_by FROM ask_egm_attachment WHERE id = :aid"
    ), {"aid": att_id})).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found")
    await _check_review_access(db, user, str(row["domain_review_id"]))

    # Only the uploader or admin/governance_lead can delete
    if not (Role.ADMIN in user.roles or Role.GOVERNANCE_LEAD in user.roles):
        if row["create_by"] != user.id:
            raise HTTPException(status_code=403, detail="Only the uploader or an admin can delete this attachment")

    await db.execute(text("DELETE FROM ask_egm_attachment WHERE id = :aid"), {"aid": att_id})
    await db.commit()
    return {"ok": True}


@router.get(
    "/{domain_review_id}/history",
    dependencies=[Depends(require_permission("review_action", "read"))],
)
async def get_history(
    domain_review_id: str,
    user: AuthUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return conversation history for a domain review."""
    await _check_review_access(db, user, domain_review_id)
    rows = (await db.execute(text("""
        SELECT id, role, content, create_by, create_at, metadata
        FROM ask_egm_conversation
        WHERE domain_review_id = :rid
        ORDER BY create_at
    """), {"rid": domain_review_id})).mappings().all()
    return {
        "data": [
            {
                "id": str(r["id"]),
                "role": r["role"],
                "content": r["content"],
                "createBy": r["create_by"],
                "createAt": r["create_at"].isoformat() if r["create_at"] else None,
                "metadata": r["metadata"],
            }
            for r in rows
        ]
    }


@router.delete(
    "/{domain_review_id}/history",
    dependencies=[Depends(require_permission("domain_review", "write"))],
)
async def clear_history(
    domain_review_id: str,
    user: AuthUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete all conversation history and attachments for a domain review.

    Restricted to admin / governance_lead / assigned domain_reviewer.
    """
    await _check_review_access(db, user, domain_review_id)
    await db.execute(text("""
        DELETE FROM ask_egm_attachment WHERE domain_review_id = :rid
    """), {"rid": domain_review_id})
    await db.execute(text("""
        DELETE FROM ask_egm_conversation WHERE domain_review_id = :rid
    """), {"rid": domain_review_id})
    await db.commit()
    return {"ok": True}


@router.post(
    "/{domain_review_id}/chat",
    dependencies=[Depends(require_permission("review_action", "read"))],
)
async def chat(
    domain_review_id: str,
    body: ChatRequest,
    request: Request,
    user: AuthUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Send a message and receive a streaming AI response (SSE).

    Supports attachments (images sent as vision content, files as text context).
    Response includes follow-up question suggestions in the 'done' event.
    """
    await _check_review_access(db, user, domain_review_id)

    user_message = body.message.strip()
    if not user_message and not body.attachmentIds:
        raise HTTPException(status_code=422, detail="Message cannot be empty")
    if not user_message:
        user_message = "(Attached files)"

    # 0. Validate LLM is configured (fail fast)
    client = _get_openai_client()

    # 1. Build system prompt with context
    system_prompt = await _build_system_prompt(db, domain_review_id, web_search=body.webSearch)

    # 1b. Auto-update embedding for current review (non-blocking, best-effort)
    try:
        await upsert_review_embedding(db, domain_review_id)
    except Exception as e:
        logger.debug("Auto-embed skipped: %s", e)
        await db.rollback()

    # 2. Fetch attachment data if provided
    attachment_data: list[dict] = []
    attachment_meta: list[dict] = []
    if body.attachmentIds:
        att_rows = (await db.execute(text("""
            SELECT id, file_name, file_size, content_type, file_data
            FROM ask_egm_attachment
            WHERE id = ANY(:ids) AND domain_review_id = :rid
        """), {"ids": body.attachmentIds, "rid": domain_review_id})).mappings().all()
        for att in att_rows:
            attachment_data.append({
                "id": str(att["id"]),
                "fileName": att["file_name"],
                "fileSize": att["file_size"],
                "contentType": att["content_type"],
                "data": bytes(att["file_data"]),
            })
            attachment_meta.append({
                "id": str(att["id"]),
                "fileName": att["file_name"],
                "contentType": att["content_type"],
            })

    # 3. Save user message with metadata
    user_metadata = None
    if attachment_meta:
        user_metadata = json.dumps({"attachments": attachment_meta})
    await db.execute(text("""
        INSERT INTO ask_egm_conversation (domain_review_id, role, content, create_by, metadata)
        VALUES (:rid, 'user', :content, :uid, CAST(:meta AS jsonb))
    """), {"rid": domain_review_id, "content": user_message, "uid": user.id, "meta": user_metadata})
    await db.commit()

    # 4. Load conversation history — replay image attachments as multimodal content
    #    so follow-up questions can still "see" previously uploaded images.
    history_rows = (await db.execute(text("""
        SELECT role, content, metadata FROM ask_egm_conversation
        WHERE domain_review_id = :rid
        ORDER BY create_at
    """), {"rid": domain_review_id})).mappings().all()

    # Collect all attachment IDs referenced in historical user messages
    history_att_ids: list[str] = []
    for h in history_rows:
        meta = h.get("metadata")
        if h["role"] == "user" and meta and isinstance(meta, dict):
            for att_ref in (meta.get("attachments") or []):
                if att_ref.get("id"):
                    history_att_ids.append(att_ref["id"])

    # Batch-fetch historical attachments for multimodal replay
    history_att_map: dict[str, dict] = {}
    if history_att_ids:
        hatt_rows = (await db.execute(text("""
            SELECT id, file_name, file_size, content_type, file_data
            FROM ask_egm_attachment WHERE id = ANY(:ids)
        """), {"ids": history_att_ids})).mappings().all()
        for att in hatt_rows:
            history_att_map[str(att["id"])] = {
                "id": str(att["id"]),
                "fileName": att["file_name"],
                "fileSize": att["file_size"],
                "contentType": att["content_type"],
                "data": bytes(att["file_data"]),
            }

    messages: list[dict] = [{"role": "system", "content": system_prompt}]
    for h in history_rows:
        meta = h.get("metadata")
        msg_atts: list[dict] = []
        if h["role"] == "user" and meta and isinstance(meta, dict):
            for att_ref in (meta.get("attachments") or []):
                att_entry = history_att_map.get(att_ref.get("id", ""))
                if att_entry:
                    msg_atts.append(att_entry)

        if msg_atts:
            messages.append({"role": h["role"], "content": _build_multimodal_content(h["content"], msg_atts)})
        else:
            messages.append({"role": h["role"], "content": h["content"]})

    # 5. For the current message, override with fresh attachment data if provided
    if attachment_data:
        multimodal = _build_multimodal_content(user_message, attachment_data)
        messages[-1] = {"role": "user", "content": multimodal}

    # 6. Stream from LLM

    async def sse_generator() -> AsyncGenerator[str, None]:
        full_response = ""
        follow_ups: list[str] = []
        search_sources: list[dict] = []
        cancelled = False

        try:
            # Prepare tool calling params if Tavily is configured
            tool_kwargs: dict = {}
            if body.webSearch and is_tavily_configured():
                tool_kwargs["tools"] = [WEB_SEARCH_TOOL]

            # ── First LLM call (may trigger tool use) ────────────────────────
            stream = await client.chat.completions.create(
                model=settings.LLM_MODEL,
                messages=messages,
                stream=True,
                temperature=settings.LLM_TEMPERATURE,
                top_p=settings.LLM_TOP_P,
                max_tokens=4096,
                **tool_kwargs,
            )

            # Accumulate streaming response — track both content and tool_calls
            tool_calls_acc: dict[int, dict] = {}  # index -> {id, name, arguments}
            finish_reason = None
            chunk_count = 0

            async for chunk in stream:
                chunk_count += 1
                if chunk_count % 10 == 0 and await request.is_disconnected():
                    cancelled = True
                    break

                if not chunk.choices:
                    continue
                delta = chunk.choices[0].delta
                finish_reason = chunk.choices[0].finish_reason or finish_reason

                # Content tokens → stream to client
                if delta.content:
                    token = delta.content
                    full_response += token
                    yield f"data: {json.dumps({'token': token})}\n\n"

                # Tool call deltas → accumulate
                if delta.tool_calls:
                    for tc_delta in delta.tool_calls:
                        idx = tc_delta.index
                        if idx not in tool_calls_acc:
                            tool_calls_acc[idx] = {
                                "id": tc_delta.id or "",
                                "name": "",
                                "arguments": "",
                            }
                        if tc_delta.id:
                            tool_calls_acc[idx]["id"] = tc_delta.id
                        if tc_delta.function:
                            if tc_delta.function.name:
                                tool_calls_acc[idx]["name"] = tc_delta.function.name
                            if tc_delta.function.arguments:
                                tool_calls_acc[idx]["arguments"] += tc_delta.function.arguments

            # ── Handle tool calls (web search) ───────────────────────────────
            if finish_reason == "tool_calls" and tool_calls_acc:
                for idx in sorted(tool_calls_acc):
                    tc = tool_calls_acc[idx]
                    if tc["name"] != "web_search":
                        continue

                    try:
                        args = json.loads(tc["arguments"])
                        query = args.get("query", "")
                    except (json.JSONDecodeError, KeyError):
                        query = ""

                    if not query:
                        continue

                    # Notify frontend: searching
                    yield f"data: {json.dumps({'searching': True, 'query': query})}\n\n"

                    # Execute Tavily search
                    results = await tavily_search(query)
                    search_sources = [{"title": r["title"], "url": r["url"]} for r in results]
                    result_text = format_search_results_for_llm(results)

                    # Notify frontend: search done with sources
                    yield f"data: {json.dumps({'searching': False, 'sources': search_sources})}\n\n"

                    # Append tool call + result to messages for second LLM call
                    messages.append({
                        "role": "assistant",
                        "content": None,
                        "tool_calls": [{
                            "id": tc["id"],
                            "type": "function",
                            "function": {
                                "name": "web_search",
                                "arguments": tc["arguments"],
                            },
                        }],
                    })
                    messages.append({
                        "role": "tool",
                        "tool_call_id": tc["id"],
                        "content": result_text,
                    })

                # ── Second LLM call: generate final answer with search context ──
                if not cancelled:
                    stream2 = await client.chat.completions.create(
                        model=settings.LLM_MODEL,
                        messages=messages,
                        stream=True,
                        temperature=settings.LLM_TEMPERATURE,
                        top_p=settings.LLM_TOP_P,
                        max_tokens=4096,
                        # No tools in second call — force text response
                    )
                    chunk_count = 0
                    async for chunk in stream2:
                        chunk_count += 1
                        if chunk_count % 10 == 0 and await request.is_disconnected():
                            cancelled = True
                            break
                        if chunk.choices and chunk.choices[0].delta.content:
                            token = chunk.choices[0].delta.content
                            full_response += token
                            yield f"data: {json.dumps({'token': token})}\n\n"

            # Extract follow-up questions from the response
            cleaned_content, follow_ups = _extract_follow_ups(full_response)
            full_response = cleaned_content

            # Signal completion
            done_payload: dict = {"done": True}
            if follow_ups:
                done_payload["followUpQuestions"] = follow_ups
            if search_sources:
                done_payload["sources"] = search_sources
            yield f"data: {json.dumps(done_payload)}\n\n"

        except Exception as e:
            logger.error("Ask EGM streaming error: %s", e, exc_info=True)
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
        except asyncio.CancelledError:
            cancelled = True
        finally:
            # Save assistant response to DB (use a new session to avoid issues)
            if full_response:
                if cancelled:
                    full_response += "\n\n*(Cancelled)*"
                try:
                    from app.database import AsyncSessionLocal
                    async with AsyncSessionLocal() as save_db:
                        asst_meta: dict = {}
                        if cancelled:
                            asst_meta["cancelled"] = True
                        if follow_ups:
                            asst_meta["followUpQuestions"] = follow_ups
                        if search_sources:
                            asst_meta["sources"] = search_sources
                        meta_json = json.dumps(asst_meta) if asst_meta else None
                        await save_db.execute(text("""
                            INSERT INTO ask_egm_conversation
                                (domain_review_id, role, content, create_by, metadata)
                            VALUES (:rid, 'assistant', :content, 'ask-egm', CAST(:meta AS jsonb))
                        """), {
                            "rid": domain_review_id,
                            "content": full_response,
                            "meta": meta_json,
                        })
                        await save_db.commit()
                except Exception as save_err:
                    logger.error("Failed to save assistant response: %s", save_err)

    return StreamingResponse(
        sse_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )

@router.post(
    "/{domain_review_id}/embedding",
    dependencies=[Depends(require_permission("review_action", "read"))],
)
async def sync_embedding(
    domain_review_id: str,
    user: AuthUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Generate/update the embedding for a domain review (for RAG)."""
    await _check_review_access(db, user, domain_review_id)
    ok = await upsert_review_embedding(db, domain_review_id)
    return {"ok": ok, "message": "Embedding updated" if ok else "No change or embedding not configured"}


@router.post(
    "/embeddings/sync-all",
    dependencies=[Depends(require_permission("domain_review", "write"))],
)
async def sync_all_embeddings(
    user: AuthUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Sync embeddings for all domain reviews that have terminal status."""
    rows = (await db.execute(text("""
        SELECT id FROM domain_review
        WHERE status IN ('Approved', 'Approved with Exception', 'Not Passed', 'Complete')
    """))).scalars().all()

    created = 0
    skipped = 0
    for rid in rows:
        try:
            ok = await upsert_review_embedding(db, str(rid))
            if ok:
                created += 1
            else:
                skipped += 1
        except Exception as e:
            logger.warning("Failed to sync embedding for %s: %s", rid, e)
            skipped += 1

    return {"total": len(rows), "created": created, "skipped": skipped}
